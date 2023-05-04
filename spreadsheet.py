import openpyxl

wb = openpyxl.load_workbook("expo_new.xlsx")
ws = wb.active

loaded_projects = []

class ProjectInfo:

    def __init__(self, name:str, desc:str, members:str, floor:int) -> None:
        self.name = name
        self.desc = desc
        self.members = members
        self.floor = floor

    def __str__(self) -> str:
        return f"(Project) {self.name}: {self.desc} [{self.members}] ({self.floor})"

print(ws['A1'].value)

# worksheet.iter_rows(min_row, max_row, min_col, max_col) generates iteration for each row in worksheet (list of cell objects that have to have their value accessed with .value)
# worksheets start their index with 1, not 0
# row 1 will be our header/labels, everything following should be projects

for row in ws.iter_rows(2):
    newProject = ProjectInfo(row[0].value, row[1].value, row[2].value, row[3].value)
    loaded_projects.append(newProject)

for project in loaded_projects:
    print(project)

def returnProjects():
    return loaded_projects